# pip --default-timeout=1000 install -U openpyxl
from openpyxl import Workbook
 
workbook = Workbook()
# 获取当前活跃的sheet,默认是第一个sheet
worksheet = workbook.active     

# 表头
worksheet.cell(1, 1, '数据列1')
worksheet.cell(1, 2, '数据列2')
worksheet.cell(1, 3, '数据列3')
worksheet.cell(1, 4, '数据列4')
worksheet.cell(1, 5, '数据列5')
worksheet.cell(1, 6, '数据列6')
worksheet.cell(1, 7, '数据列7')
worksheet.cell(1, 8, '数据列8')
worksheet.cell(1, 9, '数据列9')
worksheet.cell(1, 10, '数据列10')
worksheet.cell(1, 11, '数据列11')
worksheet.cell(1, 12, '数据列12')
worksheet.cell(1, 13, '数据列13')
worksheet.cell(1, 14, '数据列14')
worksheet.cell(1, 15, '数据列15')
worksheet.cell(1, 16, '数据列16')
worksheet.cell(1, 17, '数据列17')
worksheet.cell(1, 18, '数据列18')
worksheet.cell(1, 19, '数据列19')
worksheet.cell(1, 20, '数据列20')
worksheet.cell(1, 21, '数据列21')
# 数据
i = 1
while i < 100001:
    index = i + 1
    worksheet.cell(index, 1, '李四' + str(i))
    worksheet.cell(index, 2, 99.5)
    worksheet.cell(index, 3, '张三' + str(i))
    worksheet.cell(index, 4, 99.5)
    worksheet.cell(index, 5, 99.5)
    worksheet.cell(index, 6, 99.5)
    worksheet.cell(index, 7, 99.5)
    worksheet.cell(index, 8, 99.5)
    worksheet.cell(index, 9, 99.5)
    worksheet.cell(index, 10, 99.5)
    worksheet.cell(index, 11, 99.5)
    worksheet.cell(index, 12, 99.5)
    worksheet.cell(index, 13, 99.5)
    worksheet.cell(index, 14, 99.5)
    worksheet.cell(index, 15, 99.5)
    worksheet.cell(index, 16, 99.5)
    worksheet.cell(index, 17, 99.5)
    worksheet.cell(index, 18, 99.5)
    worksheet.cell(index, 19, 99.5)
    worksheet.cell(index, 20, 99.5)
    worksheet.cell(index, 21, '二哥' + str(i))
    i = i + 1



workbook.save("excel.xlsx")